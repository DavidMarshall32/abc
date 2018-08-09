from django.shortcuts import render, redirect
from ABCquestionnaire.forms import ValueForm#, FeedbackForm
from ABCquestionnaire.models import Value
from django.http import HttpResponse
from django.contrib.auth.models import User
from ABCquestionnaire.singlestudent import graphs
import matplotlib
matplotlib.use('Agg')
from io import BytesIO, StringIO
import six
from django.template import loader
from ABCquestionnaire.utils import render_to_pdf
from django.template.loader import get_template


def question_values(request):
    form = ValueForm()
    ABCQuestions = {'Q1':'1. Study effectively on your own in independent private study',
        'Q2':'2. Produce your best work under examination conditions',
	    'Q3':'3. Respond to questions asked by a lecturer in front of a full lecture theatre',
	    'Q4':'4. Manage your workload to meet coursework deadlines',
	    'Q5':'5. Give a presentation to a small group of fellow students',
	    'Q6':'6. Attend most taught sessions',
	    'Q7':'7. Attain good grades in your work',
	    'Q8':'8. Engage in profitable academic debate with your peers',
	    'Q9':'9. Ask lecturers questions about the material they are teaching, during a lecture',
	    'Q10':'10. Produce coursework at the required standard',
	    'Q11':'11. Write in an appropriate academic style',
	    'Q12':'12. Be on time for lectures',
	    'Q13':'13. Pass assessments at the first attempt',
	    'Q14':'14. Plan appropriate revision schedules',
	    'Q15':'15. Remain adequately motivated throughout',
	    'Q16':'16. Produce your best work in coursework assignments',
	    'Q17':'17. Attend tutorials',
      }
    return render(request,'ABCquestionnaire/index.html', {'form':form, 'ABCQuestions':ABCQuestions})


from django.core.files.base import ContentFile
import datetime
from openpyxl import load_workbook
import csv
import xlwt
import xlrd
import os
import xlsxwriter
def create_survey(request):
	global datas
	# value=Value(user=request.user)
	if request.method=='POST':
		form=ValueForm(request.POST)
		if form.is_valid():
			# request.session['Q1'] = request.POST.get('choice1')                
			answers=form.save(commit=False)
			# answers.user=request.user
			# answers.save()
			datas=[answers.choice1,answers.choice2,answers.choice3,answers.choice4,answers.choice5,answers.choice6,
				answers.choice7,answers.choice8,answers.choice9,answers.choice10,answers.choice11,answers.choice12,
				answers.choice13,answers.choice14,answers.choice15,answers.choice16,answers.choice17
			]
			return redirect('/result')
	date=datetime.datetime.now()
	date=str(date)
	for answer in datas:
		f=open('data.xlsx', 'a')
		f.write(answer+",")
	f.write(date)
	f.write("\n")
	f.close()
	exceldata=datas
	exceldata.append(date)
	if os.path.isfile('./data.xlsx')==False:    
		wb = xlsxwriter.Workbook('data.xlsx')
		ws = wb.add_worksheet()
		
		row = 0
		col = 0
		
		for i in range(1,len(exceldata)):
			ws.write(row, i-1, "V%d"%i)
		ws.write(row,len(exceldata)-1,"datetime")
		wb.close()
		wb=load_workbook("data.xlsx")
		ws=wb.worksheets[0]   
		row=ws.max_row
		for i in range(0,len(exceldata)):
			ws.cell(row=row+1,column=i+1).value=exceldata[i]
		ws.cell(row=row+1,column=len(exceldata))
		wb.save("data.xlsx")
	else:
		wb=load_workbook("data.xlsx")
		ws=wb.worksheets[0]
		row=ws.max_row
		for i in range(0,len(exceldata)):
			ws.cell(row=row+1,column=i+1).value=exceldata[i]
		wb.save("data.xlsx")
	return render(request,'ABCquestionnaire/result.html',{"datas":datas})

# def clean(self):
# 	data=self.cleaned_data['choice1']
# 	if not data:
# 		self.add_error('choice1', 'Please select an option ')
# 	return data

def submitted_info(request):
    if 'count' not in request.session:
        request.session['count'] = 0
    request.session['count'] += 1
    return render (request, 'ABCquestionnaire/result.html')

def python_code(request):   
	global datas
	global fig5
	global finalX
	global ABCX
	# form = FeedbackForm()
	fig1,fig2,fig3,fig4,fig5,Xfinal,ABCX=graphs.python(datas)
	Xfinal=[('Grd',float("{0:.3f}".format(Xfinal[0]))),('Vrb',float("{0:.3f}".format(Xfinal[1]))),('Att',float("{0:.3f}".format(Xfinal[2]))),('Std',float("{0:.3f}".format(Xfinal[3])))]
	Xfinal=sorted(Xfinal,key=lambda x:(-x[1],x[0]))
	finalX=[]
	for i in range (0,len(Xfinal)):
		finalX.append([Xfinal[i][0],"{0:.3f}".format(Xfinal[i][1])])#so every number has 3 dp
	ABCX="{0:.3f}".format(ABCX)
	template=loader.get_template('ABCquestionnaire/final.html')
	tmp1=six.StringIO()
	fig1.savefig(tmp1, format='svg', bbox_inches='tight')    
	# c1={'svg1':tmp1.getvalue()}
	tmp2=six.StringIO()
	fig2.savefig(tmp2, format='svg', bbox_inches='tight')
	# c2={'svg2':tmp2.getvalue()}
	tmp3=six.StringIO()
	fig3.savefig(tmp3, format='svg', bbox_inches='tight')
	# c3={'svg3':tmp3.getvalue()}
	tmp4=six.StringIO()
	fig4.savefig(tmp4, format='svg', bbox_inches='tight')
	# c4={'svg4':tmp4.getvalue()}
	tmp5=six.StringIO()
	fig5.savefig(tmp5, format='svg', bbox_inches='tight')
	# c5={'svg5':tmp5.getvalue()}
	c={'svg1':tmp1.getvalue(),'svg2':tmp2.getvalue(),'svg3':tmp3.getvalue(),'svg4':tmp4.getvalue(),'svg5':tmp5.getvalue(),'finalX':finalX,'ABCX':ABCX}
	return render(request, 'ABCquestionnaire/final.html', c)

def download_figs(request):
    global fig5
    tmp5=six.StringIO()
    fig5.savefig(tmp5, format='svg', bbox_inches='tight')
    return render(request, "ABCquestionnaire/download.html", {"svg5":tmp5.getvalue()})

import base64
import numpy as np
from matplotlib import pyplot as plt
def generate_view(request, *args, **kwargs):
    template=get_template('download.html')
    global fig5
    global finalX
    global ABCX
    tmp5=six.BytesIO()
    fig5.savefig(tmp5, format='png', bbox_inches='tight')
    image_base64 = base64.b64encode(tmp5.getvalue()).decode('utf-8').replace('\n', '')
    context= {
        'finalX':finalX,
        'ABCX':ABCX,
        "svg5":image_base64,
    }
    html=template.render(context)
    pdf=render_to_pdf('download.html',context)
    # figure=download_figs(request)
    if pdf:
        response=HttpResponse(pdf, content_type='application/pdf')
        filename="Information.pdf"
        content="inline; filename='%s'" %(filename)
        response['Content-Dispotition']=content
        # download=request.GET.get("download")
        # if download:
        #     content="attachment; filename='%s'" %(filename)
        # response['Content-Disposition']=content
        return response
    return HttpResponse("Not found")

from django.core.mail import send_mail
from django.conf import settings
from django.template import loader
def newsletter(request):
	global datas
	global fig5
	global finalX
	global ABCX
	global Feedback
	tmp5=six.StringIO()
	fig5.savefig(tmp5, format='svg', bbox_inches='tight')	
	# if request.method=='POST':
	# 	form=FeedbackForm(request.POST)
	# 	if form.is_valid():
	# 		text=form.save(commit=False)
	# 		Feedback=text.feedback
	# 		subject="Questionnaire feedback"
	# 		from_email=settings.EMAIL_HOST_USER
	# 		to_email=[from_email, 'gisdakisA@cardiff.ac.uk']
	# 		message=Feedback
	# 		send_mail(subject=subject, from_email=from_email, recipient_list=to_email, message=message, fail_silently=False)		
	# 		return redirect('/final2')	
	return render(request,'ABCquestionnaire/final2.html',{'svg5':tmp5.getvalue(),'finalX':finalX,'ABCX':ABCX})






